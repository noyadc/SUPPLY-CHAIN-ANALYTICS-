from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import json
import math
import os
import secrets
import sqlite3
import statistics
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "axis.db"
SECRET_PATH = ROOT / ".secret"
SECRET = SECRET_PATH.read_bytes() if SECRET_PATH.exists() else secrets.token_bytes(32)
if not SECRET_PATH.exists():
    SECRET_PATH.write_bytes(SECRET)

app = FastAPI(title="Axis Supply Chain Analytics API", version="1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ENTITY_ALIASES = {
    "sales": {
        "date": ["date", "sale_date", "order_date", "datekey"],
        "product_id": ["product_id", "productid", "sku", "product"],
        "warehouse_id": ["warehouse_id", "warehouseid", "warehouse"],
        "quantity": ["quantity", "quantity_sold", "quantitysold", "units"],
        "revenue": ["revenue", "sales", "sales_amount", "amount"],
        "cogs": ["cogs", "cost_of_goods", "cost"],
    },
    "inventory": {
        "product_id": ["product_id", "productid", "sku", "product"],
        "product_name": ["product_name", "productname", "name"],
        "category": ["category", "product_category"],
        "warehouse_id": ["warehouse_id", "warehouseid", "warehouse"],
        "current_stock": ["current_stock", "currentstock", "stock", "on_hand"],
        "safety_stock": ["safety_stock", "safetystock"],
        "reorder_point": ["reorder_point", "reorderpoint"],
        "unit_cost": ["unit_cost", "unitcost", "cost"],
        "inventory_cost": ["inventory_cost", "inventorycost", "inventory_value", "value"],
        "age_days": ["age_days", "inventory_age", "age"],
    },
    "shipments": {
        "shipment_id": ["shipment_id", "shipmentid", "id"],
        "supplier_id": ["supplier_id", "supplierid", "supplier"],
        "warehouse_id": ["warehouse_id", "warehouseid", "warehouse"],
        "origin": ["origin", "from", "source"],
        "destination": ["destination", "to"],
        "shipment_date": ["shipment_date", "shipmentdate", "dispatch_date"],
        "delivery_date": ["delivery_date", "deliverydate"],
        "delay_days": ["delay_days", "delaydays", "delay"],
        "freight_cost": ["freight_cost", "freightcost", "transportation_cost", "cost"],
        "fuel_cost": ["fuel_cost", "fuelcost"],
    },
    "purchase_orders": {
        "po_id": ["po_id", "poid", "purchase_order", "id"],
        "date": ["date", "order_date", "po_date"],
        "supplier_id": ["supplier_id", "supplierid", "supplier"],
        "product_id": ["product_id", "productid", "sku", "product"],
        "quantity_ordered": ["quantity_ordered", "quantityordered", "ordered"],
        "quantity_received": ["quantity_received", "quantityreceived", "received"],
        "lead_time": ["lead_time", "leadtime", "lead_time_days"],
        "unit_cost": ["unit_cost", "unitcost", "cost"],
        "total_cost": ["total_cost", "totalcost", "purchase_spend", "spend"],
    },
    "products": {
        "product_id": ["product_id", "productid", "sku", "id"],
        "product_name": ["product_name", "productname", "name"],
        "category": ["category", "product_category"],
        "brand": ["brand"],
    },
    "suppliers": {
        "supplier_id": ["supplier_id", "supplierid", "id"],
        "supplier_name": ["supplier_name", "suppliername", "name"],
        "region": ["region", "country"],
        "delivery_score": ["delivery_score", "delivery_performance", "on_time_delivery"],
        "quality_score": ["quality_score", "quality", "acceptance_rate"],
        "cost_score": ["cost_score", "cost_efficiency"],
        "responsiveness_score": ["responsiveness_score", "responsiveness"],
        "defect_rate": ["defect_rate", "defects"],
    },
    "warehouses": {
        "warehouse_id": ["warehouse_id", "warehouseid", "id"],
        "warehouse_name": ["warehouse_name", "warehousename", "name"],
        "city": ["city"],
        "state": ["state", "region"],
        "capacity": ["capacity", "max_capacity"],
        "picking_efficiency": ["picking_efficiency", "pick_efficiency"],
        "processing_time": ["processing_time", "order_processing_time"],
    },
}

NUMERIC_FIELDS = {
    "quantity", "revenue", "cogs", "current_stock", "safety_stock", "reorder_point",
    "unit_cost", "inventory_cost", "age_days", "delay_days", "freight_cost", "fuel_cost",
    "quantity_ordered", "quantity_received", "lead_time", "total_cost", "delivery_score",
    "quality_score", "cost_score", "responsiveness_score", "defect_rate", "capacity",
    "picking_efficiency", "processing_time",
}


def db():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def init_db():
    with db() as con:
        con.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            organization TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS datasets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS entity_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_id INTEGER NOT NULL,
            entity TEXT NOT NULL,
            rows_json TEXT NOT NULL,
            row_count INTEGER NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(dataset_id, entity),
            FOREIGN KEY(dataset_id) REFERENCES datasets(id)
        );
        """)


init_db()


class Credentials(BaseModel):
    email: str
    password: str


class RegisterInput(Credentials):
    name: str
    organization: str


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 210_000)
    return f"{base64.urlsafe_b64encode(salt).decode()}:{base64.urlsafe_b64encode(digest).decode()}"


def verify_password(password: str, stored: str) -> bool:
    salt64, digest64 = stored.split(":")
    candidate = hash_password(password, base64.urlsafe_b64decode(salt64)).split(":")[1]
    return hmac.compare_digest(candidate, digest64)


def make_token(user_id: int) -> str:
    payload = base64.urlsafe_b64encode(json.dumps({"sub": user_id, "exp": int(time.time()) + 86400 * 7}).encode()).decode().rstrip("=")
    signature = hmac.new(SECRET, payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}.{signature}"


def current_user(authorization: str | None = None):
    from fastapi import Header
    raise RuntimeError


def auth_user(authorization: str | None = __import__("fastapi").Header(default=None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Authentication required")
    token = authorization[7:]
    try:
        payload, signature = token.split(".")
        expected = hmac.new(SECRET, payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            raise ValueError()
        padded = payload + "=" * (-len(payload) % 4)
        data = json.loads(base64.urlsafe_b64decode(padded))
        if data["exp"] < time.time():
            raise ValueError()
        with db() as con:
            user = con.execute("SELECT id,name,email,organization FROM users WHERE id=?", (data["sub"],)).fetchone()
        if not user:
            raise ValueError()
        return dict(user)
    except Exception:
        raise HTTPException(401, "Invalid or expired session")


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.post("/api/auth/register")
def register(data: RegisterInput):
    if len(data.password) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")
    email = data.email.strip().lower()
    try:
        with db() as con:
            cur = con.execute(
                "INSERT INTO users(name,email,organization,password_hash,created_at) VALUES(?,?,?,?,?)",
                (data.name.strip(), email, data.organization.strip(), hash_password(data.password), datetime.utcnow().isoformat()),
            )
            user_id = cur.lastrowid
            con.execute(
                "INSERT INTO datasets(user_id,name,created_at,updated_at) VALUES(?,?,?,?)",
                (user_id, "Primary workspace", datetime.utcnow().isoformat(), datetime.utcnow().isoformat()),
            )
    except sqlite3.IntegrityError:
        raise HTTPException(409, "An account with this email already exists")
    return {"token": make_token(user_id), "user": {"id": user_id, "name": data.name, "email": email, "organization": data.organization}}


@app.post("/api/auth/login")
def login(data: Credentials):
    with db() as con:
        user = con.execute("SELECT * FROM users WHERE email=?", (data.email.strip().lower(),)).fetchone()
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(401, "Incorrect email or password")
    return {"token": make_token(user["id"]), "user": {k: user[k] for k in ("id", "name", "email", "organization")}}


@app.get("/api/me")
def me(user=Depends(auth_user)):
    return user


def clean_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def parse_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_file(content: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        values = ws.iter_rows(values_only=True)
        headers = [str(x or "") for x in next(values)]
        return [dict(zip(headers, row)) for row in values]
    raise HTTPException(400, "Only CSV and XLSX files are supported")


def normalize(entity: str, raw_rows: list[dict]) -> tuple[list[dict], list[str]]:
    aliases = ENTITY_ALIASES[entity]
    reverse = {clean_header(alias): canonical for canonical, names in aliases.items() for alias in names}
    rows, recognized = [], set()
    for raw in raw_rows:
        normalized = {}
        for key, value in raw.items():
            canonical = reverse.get(clean_header(key))
            if canonical:
                recognized.add(canonical)
                normalized[canonical] = parse_number(value) if canonical in NUMERIC_FIELDS else str(value or "").strip()
        if any(v not in ("", 0, 0.0) for v in normalized.values()):
            rows.append(normalized)
    if not rows:
        raise HTTPException(400, f"No usable {entity.replace('_', ' ')} rows were found")
    return rows, sorted(recognized)


def owned_dataset(user_id: int, dataset_id: int | None = None):
    with db() as con:
        if dataset_id:
            row = con.execute("SELECT * FROM datasets WHERE id=? AND user_id=?", (dataset_id, user_id)).fetchone()
        else:
            row = con.execute("SELECT * FROM datasets WHERE user_id=? ORDER BY id LIMIT 1", (user_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Dataset not found")
    return dict(row)


@app.get("/api/datasets")
def datasets(user=Depends(auth_user)):
    with db() as con:
        rows = con.execute("""
            SELECT d.*, COALESCE(SUM(e.row_count),0) total_rows, COUNT(e.id) sources
            FROM datasets d LEFT JOIN entity_data e ON d.id=e.dataset_id
            WHERE d.user_id=? GROUP BY d.id ORDER BY d.updated_at DESC
        """, (user["id"],)).fetchall()
    return [dict(x) for x in rows]


@app.post("/api/upload")
async def upload(
    entity: str = Form(...),
    dataset_id: int | None = Form(default=None),
    file: UploadFile = File(...),
    user=Depends(auth_user),
):
    if entity not in ENTITY_ALIASES:
        raise HTTPException(400, f"Unknown data type: {entity}")
    dataset = owned_dataset(user["id"], dataset_id)
    content = await file.read()
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(413, "File is larger than the 15 MB limit")
    raw = parse_file(content, file.filename or "upload.csv")
    rows, fields = normalize(entity, raw)
    now = datetime.utcnow().isoformat()
    with db() as con:
        con.execute("""
            INSERT INTO entity_data(dataset_id,entity,rows_json,row_count,updated_at)
            VALUES(?,?,?,?,?)
            ON CONFLICT(dataset_id,entity) DO UPDATE SET
              rows_json=excluded.rows_json,row_count=excluded.row_count,updated_at=excluded.updated_at
        """, (dataset["id"], entity, json.dumps(rows), len(rows), now))
        con.execute("UPDATE datasets SET updated_at=? WHERE id=?", (now, dataset["id"]))
    return {"entity": entity, "rows": len(rows), "recognized_fields": fields, "dataset_id": dataset["id"]}


def load_entities(dataset_id: int) -> dict[str, list[dict]]:
    with db() as con:
        records = con.execute("SELECT entity,rows_json FROM entity_data WHERE dataset_id=?", (dataset_id,)).fetchall()
    return {r["entity"]: json.loads(r["rows_json"]) for r in records}


def safe_mean(values, default=0):
    vals = [float(x) for x in values if x is not None]
    return statistics.mean(vals) if vals else default


def pct(n, d):
    return n / d * 100 if d else 0


def month_label(date_value: str) -> str:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_value[:10], fmt).strftime("%b")
        except (ValueError, TypeError):
            pass
    return str(date_value)[:7] or "Unknown"


def build_analysis(data: dict[str, list[dict]]) -> dict:
    sales = data.get("sales", [])
    inventory = data.get("inventory", [])
    shipments = data.get("shipments", [])
    pos = data.get("purchase_orders", [])
    products = {str(x.get("product_id")): x for x in data.get("products", [])}
    suppliers_source = data.get("suppliers", [])
    suppliers = {str(x.get("supplier_id")): x for x in suppliers_source}
    warehouses_source = data.get("warehouses", [])
    warehouses = {str(x.get("warehouse_id")): x for x in warehouses_source}

    revenue = sum(x.get("revenue", 0) for x in sales)
    cogs = sum(x.get("cogs", 0) for x in sales)
    if not cogs:
        cogs = revenue * .66
    gross_margin = pct(revenue - cogs, revenue)
    inventory_value = sum(x.get("inventory_cost", 0) or x.get("current_stock", 0) * x.get("unit_cost", 0) for x in inventory)
    procurement_cost = sum(x.get("total_cost", 0) or x.get("quantity_ordered", 0) * x.get("unit_cost", 0) for x in pos)
    on_time = pct(sum(1 for x in shipments if x.get("delay_days", 0) <= 0), len(shipments))
    fill_rate = pct(sum(x.get("quantity_received", 0) for x in pos), sum(x.get("quantity_ordered", 0) for x in pos))
    lead_time = safe_mean([x.get("lead_time", 0) for x in pos])
    turnover = cogs / inventory_value if inventory_value else 0

    monthly = defaultdict(lambda: {"revenue": 0, "cogs": 0})
    for row in sales:
        key = month_label(row.get("date", ""))
        monthly[key]["revenue"] += row.get("revenue", 0)
        monthly[key]["cogs"] += row.get("cogs", 0)
    revenue_trend = []
    for key, vals in monthly.items():
        rev = vals["revenue"]
        revenue_trend.append({"m": key, "revenue": round(rev / 1_000_000, 3), "target": round(rev / 1_000_000 * .94, 3), "margin": round(pct(rev - (vals["cogs"] or rev * .66), rev), 1)})

    inv_rows = []
    understock = overstock = 0
    for row in inventory:
        pid = str(row.get("product_id", ""))
        product = products.get(pid, {})
        stock, safety = row.get("current_stock", 0), row.get("safety_stock", 0)
        reorder = row.get("reorder_point", safety * 1.25)
        status = "Understock" if stock < safety else "Overstock" if stock > max(reorder * 2, safety * 2.5) else "Healthy"
        understock += status == "Understock"
        overstock += status == "Overstock"
        inv_rows.append({
            "sku": pid, "product": row.get("product_name") or product.get("product_name") or pid,
            "category": row.get("category") or product.get("category") or "Uncategorized",
            "stock": round(stock), "safety": round(safety), "reorder": round(reorder),
            "value": round(row.get("inventory_cost", 0) or stock * row.get("unit_cost", 0), 2),
            "age": round(row.get("age_days", 0)), "status": status,
        })

    warehouse_stats = defaultdict(lambda: {"stock": 0, "sales": 0})
    for x in inventory:
        warehouse_stats[str(x.get("warehouse_id", "Unknown"))]["stock"] += x.get("current_stock", 0)
    for x in sales:
        warehouse_stats[str(x.get("warehouse_id", "Unknown"))]["sales"] += x.get("quantity", 0)
    facility_rows = []
    for wid, stat in warehouse_stats.items():
        wh = warehouses.get(wid, {})
        cap = wh.get("capacity", 0) or max(stat["stock"] * 1.2, 1)
        util = min(100, pct(stat["stock"], cap))
        facility_rows.append({
            "name": wh.get("warehouse_name") or wid, "code": wid,
            "region": ", ".join(filter(None, [wh.get("city"), wh.get("state")])) or "—",
            "utilization": round(util, 1), "throughput": round(stat["sales"] / 30, 1),
            "efficiency": round(wh.get("picking_efficiency", 85), 1),
            "processing": round(wh.get("processing_time", 3.5), 1),
            "status": "Capacity risk" if util >= 90 else "Watch" if util >= 82 else "Healthy",
            "change": 0,
        })

    supplier_spend = defaultdict(float)
    supplier_lead = defaultdict(list)
    for po in pos:
        sid = str(po.get("supplier_id", "Unknown"))
        supplier_spend[sid] += po.get("total_cost", 0) or po.get("quantity_ordered", 0) * po.get("unit_cost", 0)
        supplier_lead[sid].append(po.get("lead_time", 0))
    supplier_rows = []
    all_supplier_ids = set(suppliers) | set(supplier_spend)
    for sid in all_supplier_ids:
        s = suppliers.get(sid, {})
        delivery = s.get("delivery_score", 100)
        quality = s.get("quality_score", max(0, 100 - s.get("defect_rate", 0)))
        cost = s.get("cost_score", 85)
        response = s.get("responsiveness_score", 85)
        score = delivery * .4 + quality * .3 + cost * .2 + response * .1
        supplier_rows.append({
            "name": s.get("supplier_name") or sid, "id": sid, "region": s.get("region") or "—",
            "delivery": round(delivery, 1), "quality": round(quality, 1), "cost": round(cost, 1),
            "response": round(response, 1), "spend_value": round(supplier_spend[sid], 2),
            "spend": money(supplier_spend[sid]), "score": round(score), "lead_time": round(safe_mean(supplier_lead[sid]), 1),
        })
    supplier_rows.sort(key=lambda x: x["score"], reverse=True)

    route_stats = defaultdict(lambda: {"count": 0, "ontime": 0, "days": [], "cost": 0})
    for sh in shipments:
        route = f"{sh.get('origin') or 'Origin'} → {sh.get('destination') or sh.get('warehouse_id') or 'Destination'}"
        r = route_stats[route]
        r["count"] += 1
        r["ontime"] += sh.get("delay_days", 0) <= 0
        r["days"].append(max(0, sh.get("delay_days", 0)))
        r["cost"] += sh.get("freight_cost", 0)
    routes = [{
        "route": name, "shipments": x["count"], "success": round(pct(x["ontime"], x["count"]), 1),
        "transit": round(safe_mean(x["days"]), 1), "cost": round(x["cost"] / x["count"], 2)
    } for name, x in route_stats.items()]

    daily = defaultdict(float)
    for row in sales:
        daily[str(row.get("date", ""))] += row.get("quantity", 0)
    actual = [{"d": k, "actual": round(v, 1), "forecast": None, "low": None, "high": None} for k, v in list(daily.items())[-14:]]
    base = safe_mean(list(daily.values())[-30:]) or 0
    forecast = []
    today = datetime.utcnow()
    for i in range(1, 13):
        value = base * (1 + .018 * i) * (1 + .08 * math.sin(i / 2))
        forecast.append({"d": (today + timedelta(days=i * 7)).strftime("%b %d"), "actual": None, "forecast": round(value, 1), "low": round(value * .88, 1), "high": round(value * 1.12, 1)})

    category_spend = defaultdict(float)
    for po in pos:
        product = products.get(str(po.get("product_id", "")), {})
        category_spend[product.get("category") or "Other"] += po.get("total_cost", 0) or po.get("quantity_ordered", 0) * po.get("unit_cost", 0)
    spend_total = sum(category_spend.values())
    colors = ["#3b82f6", "#9b8afb", "#22d3ee", "#35d399", "#7f91aa", "#f5b942"]
    spend = [{"name": k, "value": round(pct(v, spend_total), 1), "amount": v, "color": colors[i % len(colors)]} for i, (k, v) in enumerate(category_spend.items())]

    stockout_rate = pct(understock, len(inventory))
    avg_supplier = safe_mean([x["score"] for x in supplier_rows], 80)
    inventory_health = max(0, 100 - stockout_rate - pct(overstock, len(inventory)) * .5)
    overall = round(safe_mean([avg_supplier, inventory_health, on_time or 80, 88]))
    alerts = []
    if understock:
        alerts.append({"level": "critical", "title": f"{understock} SKUs below safety stock", "detail": "Inventory requires action", "target": "inventory"})
    if routes and min(x["success"] for x in routes) < 90:
        worst = min(routes, key=lambda x: x["success"])
        alerts.append({"level": "warning", "title": "Shipment delay risk increased", "detail": worst["route"], "target": "logistics"})
    if facility_rows and max(x["utilization"] for x in facility_rows) >= 90:
        wh = max(facility_rows, key=lambda x: x["utilization"])
        alerts.append({"level": "warning", "title": "Warehouse approaching capacity", "detail": f"{wh['name']} · {wh['utilization']}% utilized", "target": "warehouse"})

    return {
        "has_data": bool(data),
        "summary": {
            "revenue": revenue, "gross_margin": gross_margin, "inventory_value": inventory_value,
            "procurement_cost": procurement_cost, "on_time_delivery": on_time, "fill_rate": fill_rate,
            "avg_lead_time": lead_time, "inventory_turnover": turnover, "current_inventory": inventory_value,
            "stockout_rate": stockout_rate, "excess_inventory": sum(x["value"] for x in inv_rows if x["status"] == "Overstock"),
            "supplier_reliability": avg_supplier, "transportation_cost": sum(x.get("freight_cost", 0) for x in shipments),
            "shipment_count": len(shipments), "purchase_spend": procurement_cost, "active_suppliers": len(all_supplier_ids),
            "health_score": overall,
        },
        "health": [
            {"name": "Supplier", "value": round(avg_supplier), "color": "#35d399"},
            {"name": "Inventory", "value": round(inventory_health), "color": "#f5b942"},
            {"name": "Logistics", "value": round(on_time or 80), "color": "#3b82f6"},
            {"name": "Demand", "value": 88, "color": "#9b8afb"},
        ],
        "revenue_trend": revenue_trend, "inventory": inv_rows, "warehouses": facility_rows,
        "suppliers": supplier_rows, "routes": routes, "forecast": actual + forecast,
        "spend_categories": spend, "alerts": alerts,
        "sources": {k: len(v) for k, v in data.items()},
    }


def money(value):
    if value >= 1_000_000:
        return f"${value / 1_000_000:.2f}M"
    if value >= 1_000:
        return f"${value / 1_000:.0f}k"
    return f"${value:,.0f}"


@app.get("/api/analytics")
def analytics(dataset_id: int | None = None, user=Depends(auth_user)):
    dataset = owned_dataset(user["id"], dataset_id)
    result = build_analysis(load_entities(dataset["id"]))
    result["dataset"] = dataset
    return result


TEMPLATE_FIELDS = {
    "sales": ["date", "product_id", "warehouse_id", "quantity", "revenue", "cogs"],
    "inventory": ["product_id", "product_name", "category", "warehouse_id", "current_stock", "safety_stock", "reorder_point", "unit_cost", "inventory_cost", "age_days"],
    "shipments": ["shipment_id", "supplier_id", "warehouse_id", "origin", "destination", "shipment_date", "delivery_date", "delay_days", "freight_cost", "fuel_cost"],
    "purchase_orders": ["po_id", "date", "supplier_id", "product_id", "quantity_ordered", "quantity_received", "lead_time", "unit_cost", "total_cost"],
    "products": ["product_id", "product_name", "category", "brand"],
    "suppliers": ["supplier_id", "supplier_name", "region", "delivery_score", "quality_score", "cost_score", "responsiveness_score", "defect_rate"],
    "warehouses": ["warehouse_id", "warehouse_name", "city", "state", "capacity", "picking_efficiency", "processing_time"],
}


@app.get("/api/templates/{entity}.csv")
def template(entity: str):
    if entity not in TEMPLATE_FIELDS:
        raise HTTPException(404, "Template not found")
    stream = io.StringIO()
    csv.writer(stream).writerow(TEMPLATE_FIELDS[entity])
    return StreamingResponse(iter([stream.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{entity}_template.csv"'})


def demo_data():
    products = [
        {"product_id": "EL-1048", "product_name": "SmartHub Controller", "category": "Electronics", "brand": "Axis"},
        {"product_id": "HA-2201", "product_name": "AirPure Filter XL", "category": "Home appliance", "brand": "Northstar"},
        {"product_id": "IN-8834", "product_name": "Drive Belt Assembly", "category": "Industrial", "brand": "Coreline"},
        {"product_id": "PK-4412", "product_name": "Reinforced Carton L", "category": "Packaging", "brand": "Meridian"},
    ]
    warehouses = [
        {"warehouse_id": "CHI-01", "warehouse_name": "Chicago Central", "city": "Chicago", "state": "IL", "capacity": 4500, "picking_efficiency": 83, "processing_time": 3.8},
        {"warehouse_id": "DAL-02", "warehouse_name": "Dallas Fulfillment", "city": "Dallas", "state": "TX", "capacity": 3900, "picking_efficiency": 94, "processing_time": 2.7},
    ]
    suppliers = [
        {"supplier_id": "SUP-01", "supplier_name": "Nexus Components", "region": "Taiwan", "delivery_score": 96, "quality_score": 94, "cost_score": 87, "responsiveness_score": 92},
        {"supplier_id": "SUP-02", "supplier_name": "Meridian Pack", "region": "Mexico", "delivery_score": 72, "quality_score": 79, "cost_score": 91, "responsiveness_score": 68},
    ]
    inventory = [
        {"product_id": "EL-1048", "warehouse_id": "CHI-01", "current_stock": 118, "safety_stock": 160, "reorder_point": 190, "unit_cost": 1210, "inventory_cost": 142780, "age_days": 18},
        {"product_id": "HA-2201", "warehouse_id": "CHI-01", "current_stock": 842, "safety_stock": 310, "reorder_point": 390, "unit_cost": 140, "inventory_cost": 117880, "age_days": 127},
        {"product_id": "IN-8834", "warehouse_id": "DAL-02", "current_stock": 64, "safety_stock": 85, "reorder_point": 105, "unit_cost": 570, "inventory_cost": 36480, "age_days": 42},
        {"product_id": "PK-4412", "warehouse_id": "DAL-02", "current_stock": 2860, "safety_stock": 1200, "reorder_point": 1500, "unit_cost": 20, "inventory_cost": 57200, "age_days": 16},
    ]
    sales, pos, shipments = [], [], []
    today = datetime.utcnow()
    for i in range(180):
        day = today - timedelta(days=179-i)
        pid = products[i % len(products)]["product_id"]
        wid = warehouses[i % len(warehouses)]["warehouse_id"]
        qty = 110 + (i * 17) % 90
        sales.append({"date": day.strftime("%Y-%m-%d"), "product_id": pid, "warehouse_id": wid, "quantity": qty, "revenue": qty * (180 + (i % 4) * 90), "cogs": qty * (112 + (i % 4) * 57)})
    for i in range(36):
        pos.append({"po_id": f"PO-{i+1:04}", "date": (today-timedelta(days=i*8)).strftime("%Y-%m-%d"), "supplier_id": suppliers[i%2]["supplier_id"], "product_id": products[i%4]["product_id"], "quantity_ordered": 500+i*10, "quantity_received": 480+i*10 if i%5 else 420+i*10, "lead_time": 11+i%8, "unit_cost": 95+i%4*70, "total_cost": (500+i*10)*(95+i%4*70)})
    for i in range(80):
        delay = -1 if i % 9 else 4
        shipments.append({"shipment_id": f"SHP-{i+1:04}", "supplier_id": suppliers[i%2]["supplier_id"], "warehouse_id": warehouses[i%2]["warehouse_id"], "origin": "Taipei" if i%2==0 else "Monterrey", "destination": "Chicago" if i%2==0 else "Dallas", "shipment_date": (today-timedelta(days=i*2+12)).strftime("%Y-%m-%d"), "delivery_date": (today-timedelta(days=i*2)).strftime("%Y-%m-%d"), "delay_days": delay, "freight_cost": 820+i%7*55, "fuel_cost": 130+i%5*10})
    return {"products": products, "warehouses": warehouses, "suppliers": suppliers, "inventory": inventory, "sales": sales, "purchase_orders": pos, "shipments": shipments}


@app.post("/api/demo")
def load_demo(dataset_id: int | None = None, user=Depends(auth_user)):
    dataset = owned_dataset(user["id"], dataset_id)
    now = datetime.utcnow().isoformat()
    with db() as con:
        for entity, rows in demo_data().items():
            con.execute("""
                INSERT INTO entity_data(dataset_id,entity,rows_json,row_count,updated_at) VALUES(?,?,?,?,?)
                ON CONFLICT(dataset_id,entity) DO UPDATE SET rows_json=excluded.rows_json,row_count=excluded.row_count,updated_at=excluded.updated_at
            """, (dataset["id"], entity, json.dumps(rows), len(rows), now))
        con.execute("UPDATE datasets SET updated_at=? WHERE id=?", (now, dataset["id"]))
    return {"message": "Demo dataset loaded", "dataset_id": dataset["id"]}
